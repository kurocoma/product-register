"""
extract_test_data.py

Reads the Excel template and extracts:
1. First 5 products from データ入力シート -> tests/fixtures/input_sample.csv
2. Matching rows from 楽天用商品登録データ -> tests/fixtures/expected/rakuten_normal_item.csv
3. Matching rows from 商品ページ生成_単品商品 -> tests/fixtures/expected/ne_single.csv
4. Matching rows from 商品ページ生成_セット商品 -> tests/fixtures/expected/ne_set.csv
5. Matching rows from shopify -> tests/fixtures/expected/shopify.csv

Usage:
    python scripts/extract_test_data.py
"""
import csv
import sys
from pathlib import Path

import openpyxl

EXCEL_PATH = r"C:\Users\hppym\OneDrive\デスクトップ\ジャイアンツ・お菓子【最新】【重要】商品登録用テンプレート_260127.xlsm"

# Mapping from Japanese Excel headers -> ProductInput field names
HEADER_MAP = {
    "NEコード": "ne_code",
    "JANコード": "jan_code",
    "メーカーコード": "maker_code",
    "単品orセット商品": "product_type",
    "数量": "quantity",
    "商品名": "product_name",
    "掲載商品名": "display_name",
    "消費税率": "tax_rate",
    "原価": "cost_price",
    "販売価格": "selling_price",
    "送料区分": "shipping_type",
    "作成した画像数": "image_count",
    "配送方法セット": "delivery_method",
    "納期管理番号": "lead_time",
    "モール基本カテゴリID": "mall_category_id",
    "店舗内カテゴリ1": "store_category",
    "キャッチコピーPC": "catch_copy_pc",
    "キャッチコピー(Yahoo)": "catch_copy_yahoo",
    "説明文1（PC）": "description_pc",
    "説明文2（スマホ）": "description_sp",
    "説明文4(販売説明文)": "description_4",
    "free1": "free1",
    "free2": "free2",
    "キーワード": "keyword",
    "メーカー名": "maker_name",
    "ブランド名": "brand_name",
    "項目選択肢項目名": "option_item_name",
    "項目選択肢横(名前)": "option_horizontal",
    "バリエーション項目キー定義": "variation_key",
    "バリエーション項目名定義": "variation_name",
    "バリエーション1選択肢定義": "variation_choices",
    "選択肢番号一括入力欄": "choice_numbers",
    "gazou_url1": "image_url_1",
    "gazou_url2": "image_url_2",
    "gazou_url3": "image_url_3",
    "gazou_url4": "image_url_4",
    "gazou_url5": "image_url_5",
    "gazou_url6": "image_url_6",
    "gazou_url7": "image_url_7",
    "gazou_url8": "image_url_8",
    "gazou_url9": "image_url_9",
    "gazou_url10": "image_url_10",
    "gazou_url11": "image_url_11",
    "gazou_url12": "image_url_12",
    "gazou_url13": "image_url_13",
    "gazou_url14": "image_url_14",
    "gazou_url15": "image_url_15",
    "gazou_url16": "image_url_16",
    "gazou_url17": "image_url_17",
    "gazou_url18": "image_url_18",
    "gazou_url19": "image_url_19",
    "gazou_url20": "image_url_20",
    "商品属性（項目）1": "attribute_item_1",
    "商品属性（値）1": "attribute_value_1",
    "商品属性（単位）1": "attribute_unit_1",
    "商品属性（項目）2": "attribute_item_2",
    "商品属性（値）2": "attribute_value_2",
    "商品属性（単位）2": "attribute_unit_2",
    "商品属性（項目）3": "attribute_item_3",
    "商品属性（値）3": "attribute_value_3",
    "商品属性（単位）3": "attribute_unit_3",
    "商品属性（項目）4": "attribute_item_4",
    "商品属性（値）4": "attribute_value_4",
    "商品属性（単位）4": "attribute_unit_4",
    "商品属性（項目）5": "attribute_item_5",
    "商品属性（値）5": "attribute_value_5",
    "商品属性（単位）5": "attribute_unit_5",
}

# ProductInput fields that must be numeric integers
INT_FIELDS = {
    "quantity", "tax_rate", "cost_price", "selling_price",
    "image_count", "delivery_method", "lead_time",
}

FIXTURES_DIR = Path(__file__).parent.parent / "tests" / "fixtures"
EXPECTED_DIR = FIXTURES_DIR / "expected"


def cell_str(cell_value) -> str:
    """Convert a cell value to string, returning '' for None/errors."""
    if cell_value is None:
        return ""
    s = str(cell_value)
    # Skip formula error values
    if s.startswith("#"):
        return ""
    return s.strip()


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        print(f"  WARNING: No rows to write to {path}")
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys())
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Wrote {len(rows)} rows -> {path}")


def extract_input_sample(wb) -> list[dict]:
    """Extract first 5 products from データ入力シート."""
    ws = wb["データ入力シート"]

    # Row 3 = Japanese headers (1-indexed)
    headers_row = 3
    data_start_row = 4

    # Build column index -> mapped field name
    col_to_field: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        jp_header = cell_str(ws.cell(headers_row, col).value)
        if jp_header in HEADER_MAP:
            col_to_field[col] = HEADER_MAP[jp_header]

    products = []
    for row in range(data_start_row, ws.max_row + 1):
        # key column is NEコード (col index found for 'ne_code')
        ne_col = next((c for c, f in col_to_field.items() if f == "ne_code"), None)
        if ne_col is None:
            break
        ne_val = cell_str(ws.cell(row, ne_col).value)
        if not ne_val:
            continue  # skip empty rows

        record: dict[str, str] = {}
        for col, field in col_to_field.items():
            val = cell_str(ws.cell(row, col).value)
            # For numeric fields, ensure empty becomes "0" or keep as-is
            if field in INT_FIELDS:
                if val == "" or val == "0":
                    record[field] = "0"
                else:
                    record[field] = val
            else:
                record[field] = val

        products.append(record)
        if len(products) >= 5:
            break

    return products


def _ne_base_code(ne_code: str) -> str:
    """Strip quantity suffix: t002-2542-1 -> t002-2542."""
    parts = ne_code.rsplit("-", 1)
    if len(parts) == 2:
        return parts[0]
    return ne_code


def extract_rakuten_expected(wb, ne_codes: list[str]) -> list[dict]:
    """Extract matching rows from 楽天用商品登録データ for the given NE codes."""
    ws = wb["楽天用商品登録データ"]

    # Row 1 = headers
    headers = [cell_str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    # Filter out trailing empty headers
    while headers and headers[-1] == "":
        headers.pop()

    # Build base codes from ne_codes (t002-2542-1 -> t002-2542)
    base_codes = {_ne_base_code(nc) for nc in ne_codes}

    rows = []
    for row in range(2, ws.max_row + 1):
        key_val = cell_str(ws.cell(row, 1).value)
        if not key_val:
            continue
        if key_val in base_codes:
            record = {}
            for i, h in enumerate(headers, start=1):
                record[h] = cell_str(ws.cell(row, i).value)
            rows.append(record)

    return rows


def extract_ne_single_expected(wb, ne_codes: list[str]) -> list[dict]:
    """Extract matching rows from 商品ページ生成_単品商品."""
    ws = wb["商品ページ生成_単品商品"]

    # Row 3 = API keys (field names)
    # Row 4+ = data
    api_keys_row = 3
    data_start_row = 4

    col_to_key: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        key = cell_str(ws.cell(api_keys_row, col).value)
        if key:
            col_to_key[col] = key

    # syohin_code column
    code_col = next((c for c, k in col_to_key.items() if k == "syohin_code"), None)

    rows = []
    for row in range(data_start_row, ws.max_row + 1):
        if code_col is None:
            break
        code_val = cell_str(ws.cell(row, code_col).value)
        if not code_val or code_val.startswith("#"):
            continue
        if code_val in ne_codes:
            record = {}
            for col, key in col_to_key.items():
                record[key] = cell_str(ws.cell(row, col).value)
            rows.append(record)

    return rows


def extract_ne_set_expected(wb, ne_codes: list[str]) -> list[dict]:
    """Extract matching rows from 商品ページ生成_セット商品."""
    ws = wb["商品ページ生成_セット商品"]

    # Row 3 = API keys
    # Row 4+ = data
    api_keys_row = 3
    data_start_row = 4

    col_to_key: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        key = cell_str(ws.cell(api_keys_row, col).value)
        if key:
            col_to_key[col] = key

    # set_syohin_code column
    code_col = next((c for c, k in col_to_key.items() if k == "set_syohin_code"), None)

    rows = []
    for row in range(data_start_row, ws.max_row + 1):
        if code_col is None:
            break
        code_val = cell_str(ws.cell(row, code_col).value)
        if not code_val or code_val.startswith("#"):
            continue
        if code_val in ne_codes:
            record = {}
            for col, key in col_to_key.items():
                record[key] = cell_str(ws.cell(row, col).value)
            rows.append(record)

    return rows


def extract_shopify_expected(wb, ne_codes: list[str]) -> list[dict]:
    """Extract matching rows from shopify sheet."""
    ws = wb["shopify"]

    # Row 3 = headers (skip col A which is None)
    headers_row = 3
    data_start_row = 4

    col_to_header: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        h = cell_str(ws.cell(headers_row, col).value)
        if h:
            col_to_header[col] = h

    # Handle column
    handle_col = next((c for c, h in col_to_header.items() if h == "Handle"), None)

    rows = []
    for row in range(data_start_row, ws.max_row + 1):
        if handle_col is None:
            break
        handle_val = cell_str(ws.cell(row, handle_col).value)
        if not handle_val:
            continue
        # Shopify Handle is the base code (without quantity suffix)
        # Match if handle_val == base code of any ne_code in our list
        base_codes = {_ne_base_code(nc) for nc in ne_codes}
        # Also check direct match (some ne_codes may already be base)
        all_matches = base_codes | set(ne_codes)
        if handle_val in all_matches:
            record = {}
            for col, h in col_to_header.items():
                record[h] = cell_str(ws.cell(row, col).value)
            rows.append(record)

    return rows


def main() -> None:
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    print("\n[1] Extracting input_sample.csv from データ入力シート...")
    products = extract_input_sample(wb)
    ne_codes = [p["ne_code"] for p in products]
    print(f"  Found {len(products)} products: {ne_codes}")

    # Ensure all required ProductInput fields are present with defaults
    all_fields = list(HEADER_MAP.values())
    for p in products:
        for f in all_fields:
            if f not in p:
                p[f] = "0" if f in INT_FIELDS else ""

    write_csv(FIXTURES_DIR / "input_sample.csv", products)

    print("\n[2] Extracting rakuten_normal_item.csv...")
    rakuten_rows = extract_rakuten_expected(wb, ne_codes)
    print(f"  Found {len(rakuten_rows)} rows for base codes {[_ne_base_code(c) for c in ne_codes]}")
    write_csv(EXPECTED_DIR / "rakuten_normal_item.csv", rakuten_rows)

    print("\n[3] Extracting ne_single.csv...")
    ne_single_rows = extract_ne_single_expected(wb, ne_codes)
    print(f"  Found {len(ne_single_rows)} single rows")
    write_csv(EXPECTED_DIR / "ne_single.csv", ne_single_rows)

    print("\n[4] Extracting ne_set.csv...")
    ne_set_rows = extract_ne_set_expected(wb, ne_codes)
    print(f"  Found {len(ne_set_rows)} set rows")
    write_csv(EXPECTED_DIR / "ne_set.csv", ne_set_rows)

    print("\n[5] Extracting shopify.csv...")
    shopify_rows = extract_shopify_expected(wb, ne_codes)
    print(f"  Found {len(shopify_rows)} shopify rows")
    write_csv(EXPECTED_DIR / "shopify.csv", shopify_rows)

    print("\nDone.")


if __name__ == "__main__":
    main()
