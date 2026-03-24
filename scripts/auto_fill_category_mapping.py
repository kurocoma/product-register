"""
楽天ジャンル名 → Yahoo カテゴリ ID 自動推定スクリプト

未マッチの86行について、楽天ジャンル名とYahooカテゴリ名のファジーマッチングを行い、
最適なYahooカテゴリを推定する。
"""

import csv
import re
import unicodedata
from pathlib import Path
from collections import defaultdict

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
MAPPING_CSV = BASE_DIR / "config" / "master" / "category_mapping.csv"
YAHOO_CSV = BASE_DIR / "docs" / "sampledata" / "Yahooプロダクトカテゴリ一覧data.csv"
OUTPUT_CSV = BASE_DIR / "config" / "master" / "category_mapping.csv"
OUTPUT_XLSX = BASE_DIR / "config" / "master" / "category_mapping.xlsx"

# ---------------------------------------------------------------------------
# Domain mapping: Rakuten top-level → Yahoo top-level keywords
# ---------------------------------------------------------------------------
DOMAIN_MAP = {
    "食品": ["食品"],
    "スイーツ・お菓子": ["食品"],
    "水・ソフトドリンク": ["食品"],
    "日本酒・焼酎": ["食品"],
    "ビール・洋酒": ["食品"],
    "ダイエット・健康": ["ダイエット、健康"],
    "美容・コスメ・香水": ["コスメ、美容、ヘアケア"],
    "日用品雑貨・文房具・手芸": ["キッチン、日用品、文具", "日用品、生活雑貨"],
    "インナー・下着・ナイトウェア": ["ファッション"],
    "インテリア・寝具・収納": ["家具、インテリア", "DIY、工具、ガーデン"],
}

# ---------------------------------------------------------------------------
# Manual overrides for cases that fuzzy matching cannot handle well
# ---------------------------------------------------------------------------
MANUAL_OVERRIDES: dict[str, tuple[str, str, str] | None] = {
    # (yahoo_id, yahoo_name, yahoo_path_example) or None to skip (auto)
    "日用品雑貨・文房具・手芸>日用消耗品>消臭剤・芳香剤":
        ("3986", "部屋用（芳香剤、消臭剤）", "キッチン、日用品、文具 > 芳香剤、消臭剤、除湿剤 > 部屋用"),
    "ビール・洋酒>ワイン>スパークリングワイン・シャンパン":
        ("41841", "シャンパン・スパークリングワイン", "食品 > ドリンク、水、お酒 > ワイン > シャンパン、スパークリング"),
    "ビール・洋酒>リキュール>ナッツ・種子・核系>コーヒー":
        ("44429", "リキュール", "食品 > ドリンク、水、お酒 > リキュール"),
    "ダイエット・健康>健康食品>栄養・健康ドリンク>その他":
        ("1889", "栄養ドリンク、美容健康飲料", "ダイエット、健康 > 健康飲料 > 栄養ドリンク、美容健康飲料"),
    "ビール・洋酒>リキュール>ハーブ・スパイス・ティー系>ハーブ・スパイス系>その他":
        ("44429", "リキュール", "食品 > ドリンク、水、お酒 > リキュール"),
    "食品>調味料>スパイス>唐辛子>島とうがらし":
        ("21310", "島唐辛子", "食品 > 調味料、料理の素、油 > 香辛料、スパイス、ドライハーブ > 島唐辛子"),
    "ダイエット・健康>サプリメント>動物性エキス>スッポン":
        ("1916", "ローヤルゼリー", "ダイエット、健康 > サプリメント > ローヤルゼリー"),  # closest supplement
    "ダイエット・健康>リラックス・マッサージ用品>腰ケア用品":
        ("2090", "腰痛ベルト、コルセット", "ダイエット、健康 > 矯正用品、補助ベルト > 腰痛ベルト、コルセット"),
    "ビール・洋酒>ウイスキー>ジャパニーズ・ウイスキー":
        ("1343", "国産ウイスキー", "食品 > ドリンク、水、お酒 > ウィスキー > 国産ウイスキー"),
    "食品>缶詰>肉加工品":
        ("44440", "缶詰", "食品 > 乾物、乾燥豆類、缶詰 > 缶詰"),
    "食品>缶詰>セット・詰め合わせ":
        ("44440", "缶詰", "食品 > 乾物、乾燥豆類、缶詰 > 缶詰"),
    "美容・コスメ・香水>ヘアケア・スタイリング>育毛・養毛剤":
        ("5010", "育毛、スカルプケア", "コスメ、美容、ヘアケア > ヘアケア > 育毛、スカルプケア"),
    "日用品雑貨・文房具・手芸>日用消耗品>洗剤・柔軟剤・クリーナー>カビ取り剤":
        ("43502", "防カビ洗剤", "キッチン、日用品、文具 > 掃除用具 > 洗剤 > 防カビ洗剤"),
    "スイーツ・お菓子>あめ・ミント・ガム>あめ・キャンディ":
        ("41458", "飴、ソフトキャンディ", "食品 > スナック、お菓子、おつまみ > 飴、ソフトキャンディ"),
    "水・ソフトドリンク>水・炭酸水":
        ("19033", "ミネラルウォーター、水", "食品 > ドリンク、水、お酒 > 水、炭酸水 > ミネラルウォーター、水"),
    "美容・コスメ・香水>スキンケア>フェイスクリーム":
        ("1752", "スキンケアクリーム", "コスメ、美容、ヘアケア > スキンケア > スキンケアクリーム"),
    "日本酒・焼酎>焼酎>いも焼酎":
        ("41677", "芋焼酎", "食品 > ドリンク、水、お酒 > 焼酎 > 芋焼酎"),
    "スイーツ・お菓子>スナック菓子>ポップコーン":
        ("13451", "その他スナック、お菓子、おつまみ", "食品 > スナック、お菓子、おつまみ > その他スナック、お菓子、おつまみ"),
    "美容・コスメ・香水>ボディケア>ハンドクリーム":
        ("1843", "ハンドケア用品", "コスメ、美容、ヘアケア > ボディケア > ハンドケア"),
    "スイーツ・お菓子>せんべい・米菓>あられ・おかき":
        ("1145", "あられ、ひなあられ", "食品 > 和菓子、中華菓子 > あられ、ひなあられ"),
    "ダイエット・健康>健康食品>栄養・健康ドリンク>栄養ドリンク剤":
        ("1889", "栄養ドリンク、美容健康飲料", "ダイエット、健康 > 健康飲料 > 栄養ドリンク、美容健康飲料"),
    "水・ソフトドリンク>生姜湯":
        ("44430", "ハーブティー", "食品 > ドリンク、水、お酒 > ハーブティー"),
    "食品>調味料>ドレッシング>和風ドレッシング":
        ("41572", "調味料 ドレッシング", "食品 > 調味料、料理の素、油 > ドレッシング"),
    "食品>調味料>ドレッシング>セット・詰め合わせ":
        ("41572", "調味料 ドレッシング", "食品 > 調味料、料理の素、油 > ドレッシング"),
    "食品>調味料>みそ>セット・詰め合わせ":
        ("44421", "みそ", "食品 > 調味料、料理の素、油 > みそ"),
    "スイーツ・お菓子>和菓子>落雁・干菓子":
        ("41455", "干菓子、落雁", "食品 > 和菓子、中華菓子 > 干菓子、落雁"),
    "美容・コスメ・香水>メイク道具・ケアグッズ>眉・鼻毛・甘皮はさみ":
        ("1787", "メイクブラシ", "コスメ、美容、ヘアケア > メイク道具 > メイクブラシ"),
    "スイーツ・お菓子>スナック菓子>ポテトチップス":
        ("13451", "その他スナック、お菓子、おつまみ", "食品 > スナック、お菓子、おつまみ > その他スナック、お菓子、おつまみ"),
    "美容・コスメ・香水>ヘアケア・スタイリング>ブラシ・くし":
        ("5010", "育毛、スカルプケア", "コスメ、美容、ヘアケア > ヘアケア > 育毛、スカルプケア"),
    "日用品雑貨・文房具・手芸>日用消耗品>洗剤・柔軟剤・クリーナー>浴室・浴槽洗剤":
        ("3979", "浴室洗剤", "キッチン、日用品、文具 > 掃除用具 > 洗剤 > 浴室洗剤"),
    "ダイエット・健康>サプリメント>植物性エキス>ニンニク":
        ("39452", "にんにく卵黄", "ダイエット、健康 > サプリメント > にんにく卵黄"),
    "食品>精肉・肉加工品>豚肉>モモ":
        ("1029", "豚肉、豚ホルモン", "食品 > 肉、ハム、ソーセージ > 豚肉、豚ホルモン"),
    "食品>精肉・肉加工品>豚肉>その他":
        ("1029", "豚肉、豚ホルモン", "食品 > 肉、ハム、ソーセージ > 豚肉、豚ホルモン"),
    "水・ソフトドリンク>お茶・紅茶>茶葉・ティーバッグ>植物茶":
        ("44430", "ハーブティー", "食品 > ドリンク、水、お酒 > ハーブティー"),
    "ビール・洋酒>ビール・発泡酒>ノンアルコール":
        ("19543", "ノンアルコールビール", "食品 > ドリンク、水、お酒 > ビール、発泡酒 > ノンアルコールビール"),
    "食品>精肉・肉加工品>豚肉>ロース":
        ("1029", "豚肉、豚ホルモン", "食品 > 肉、ハム、ソーセージ > 豚肉、豚ホルモン"),
    "美容・コスメ・香水>ボディケア>石けん・ボディソープ":
        ("1865", "ボディソープ", "コスメ、美容、ヘアケア > ボディケア > ボディソープ"),
    "美容・コスメ・香水>アロマ・お香>エッセンシャルオイル・精油":
        ("1933", "エッセンシャルオイル", "ダイエット、健康 > アロマグッズ > エッセンシャルオイル"),
    "ビール・洋酒>リキュール>フルーツ系>トロピカル系>マンゴー":
        ("44429", "リキュール", "食品 > ドリンク、水、お酒 > リキュール"),
    "ビール・洋酒>リキュール>ハーブ・スパイス・ティー系>ハーブ・スパイス系>ミント":
        ("44429", "リキュール", "食品 > ドリンク、水、お酒 > リキュール"),
    "ビール・洋酒>リキュール>その他":
        ("44429", "リキュール", "食品 > ドリンク、水、お酒 > リキュール"),
    "ビール・洋酒>リキュール>フルーツ系>果皮系>その他":
        ("44429", "リキュール", "食品 > ドリンク、水、お酒 > リキュール"),
    "ビール・洋酒>リキュール>フルーツ系>トロピカル系>その他":
        ("44429", "リキュール", "食品 > ドリンク、水、お酒 > リキュール"),
    "ビール・洋酒>リキュール>ハーブ・スパイス・ティー系>その他":
        ("44429", "リキュール", "食品 > ドリンク、水、お酒 > リキュール"),
    "ビール・洋酒>リキュール>フルーツ系>果皮系>ゆず":
        ("44429", "リキュール", "食品 > ドリンク、水、お酒 > リキュール"),
    "食品>惣菜>洋風惣菜>その他":
        ("1079", "その他惣菜、料理", "食品 > 惣菜、料理 > その他惣菜、料理"),
    "食品>惣菜>カレー":
        ("14643", "カレー、レトルトカレー", "食品 > 惣菜、料理 > カレー、ハヤシライス > カレー、 レトルトカレー"),
    "スイーツ・お菓子>和菓子>どら焼き>その他":
        ("1173", "どら焼き", "食品 > 和菓子、中華菓子 > どら焼き"),
    "スイーツ・お菓子>和菓子>その他":
        ("1181", "その他和菓子、中華菓子", "食品 > 和菓子、中華菓子 > その他和菓子、中華菓子"),
    "スイーツ・お菓子>製菓・製パン材料>粉類・ケーキミックス>その他":
        ("1333", "お菓子、ホットケーキミックス粉", "食品 > 米、雑穀、粉類 > お菓子、ホットケーキミックス"),
    "ダイエット・健康>健康食品>栄養調整食品>その他":
        ("1889", "栄養ドリンク、美容健康飲料", "ダイエット、健康 > 健康飲料 > 栄養ドリンク、美容健康飲料"),
    "食品>調味料>ポン酢>その他":
        ("41523", "ポン酢", "食品 > 調味料、料理の素、油 > ポン酢"),
    "インナー・下着・ナイトウェア>レディース>靴下・レッグウェア>靴下":
        ("1669", "その他レディースレッグウエア", "ファッション > レディースファッション > 下着、靴下、部屋着 > その他レッグウエア"),
    "水・ソフトドリンク>お茶・紅茶>茶葉・ティーバッグ>その他":
        ("1386", "ティーバッグ紅茶", "食品 > ドリンク、水、お酒 > 紅茶 > ティーバッグ紅茶"),
    "水・ソフトドリンク>お茶・紅茶>お茶飲料>中国茶":
        ("1397", "その他中国茶", "食品 > ドリンク、水、お酒 > 中国茶 > その他中国茶"),
    "水・ソフトドリンク>お茶・紅茶>茶葉・ティーバッグ>中国茶":
        ("1397", "その他中国茶", "食品 > ドリンク、水、お酒 > 中国茶 > その他中国茶"),
    "食品>調味料>ポン酢>ゆず":
        ("41523", "ポン酢", "食品 > 調味料、料理の素、油 > ポン酢"),
    "スイーツ・お菓子>ドライフルーツ>パイナップル":
        ("44494", "ドライフルーツ", "食品 > フルーツ > ドライフルーツ"),
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize(text: str) -> str:
    """Normalize Japanese text for comparison."""
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("　", " ").strip()
    return text


def load_yahoo_categories(path: Path) -> list[dict]:
    """Load Yahoo category master (cp932)."""
    rows = []
    with open(path, encoding="cp932", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            row["id"] = row["id"].strip()
            row["name"] = row["name"].strip()
            row["path_name"] = row["path_name"].strip()
            row["_name_norm"] = normalize(row["name"])
            row["_path_norm"] = normalize(row["path_name"])
            parts = [p.strip() for p in row["path_name"].split(">")]
            row["_top_level"] = parts[0] if parts else ""
            row["_path_parts"] = parts
            rows.append(row)
    return rows


def load_mapping(path: Path) -> list[dict]:
    """Load current mapping CSV (utf-8-sig)."""
    rows = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def parse_rakuten_genre(genre_name: str) -> list[str]:
    """Split Rakuten genre name into hierarchy levels."""
    return [p.strip() for p in genre_name.split(">")]


def get_yahoo_domain_keywords(rakuten_top: str) -> list[str]:
    """Get expected Yahoo top-level domain keywords for a Rakuten top category."""
    for key, vals in DOMAIN_MAP.items():
        if key in rakuten_top or rakuten_top in key:
            return vals
    return []


def domain_matches(yahoo_top: str, domain_keywords: list[str]) -> bool:
    """Check if Yahoo top-level category matches expected domain."""
    for kw in domain_keywords:
        if kw in yahoo_top or yahoo_top in kw:
            return True
    return False


def score_match(rakuten_parts: list[str], yahoo_row: dict, domain_keywords: list[str]) -> tuple[float, str]:
    """
    Score how well a Yahoo category matches a Rakuten genre.
    Returns (score, reasoning).
    """
    leaf = normalize(rakuten_parts[-1])
    yahoo_name = yahoo_row["_name_norm"]
    yahoo_path = yahoo_row["_path_norm"]
    yahoo_top = yahoo_row["_top_level"]

    score = 0.0
    reasons = []

    # --- Exact leaf match on Yahoo name ---
    if leaf == yahoo_name:
        score += 100
        reasons.append(f"リーフ完全一致: '{leaf}'")
    elif leaf in yahoo_name:
        ratio = len(leaf) / len(yahoo_name) if yahoo_name else 0
        # Only count if substantial overlap
        if ratio >= 0.3:
            score += 60 * ratio
            reasons.append(f"リーフがYahoo名に含まれる: '{leaf}' in '{yahoo_name}' ({ratio:.0%})")
    elif yahoo_name in leaf:
        ratio = len(yahoo_name) / len(leaf) if leaf else 0
        if ratio >= 0.3:
            score += 50 * ratio
            reasons.append(f"Yahoo名がリーフに含まれる: '{yahoo_name}' in '{leaf}' ({ratio:.0%})")

    # --- Leaf appears in path (but not in name) ---
    if leaf in yahoo_path and leaf not in yahoo_name:
        score += 30
        reasons.append(f"リーフがパスに含まれる")

    # --- Parent context matching ---
    for i in range(len(rakuten_parts) - 2, -1, -1):
        parent = normalize(rakuten_parts[i])
        if parent in yahoo_path:
            bonus = 20 if i == len(rakuten_parts) - 2 else 10
            score += bonus
            reasons.append(f"親カテゴリ '{parent}' がパスに含まれる(+{bonus})")
            break

    # --- Domain matching ---
    if domain_keywords:
        if domain_matches(yahoo_top, domain_keywords):
            score += 25
            reasons.append(f"ドメイン一致: '{yahoo_top}'")
        else:
            score -= 30
            reasons.append(f"ドメイン不一致: Yahoo='{yahoo_top}' (-30)")

    # --- Penalize overly generic categories ---
    if "その他" in yahoo_name:
        score -= 5
        reasons.append("Yahoo名に'その他'含む(-5)")

    # --- Bonus for path depth similarity ---
    r_depth = len(rakuten_parts)
    y_depth = len(yahoo_row["_path_parts"])
    if abs(r_depth - y_depth) <= 1:
        score += 5
        reasons.append("階層深さ近似")

    return score, "; ".join(reasons)


def find_best_match(rakuten_genre_name: str, yahoo_cats: list[dict],
                    yahoo_name_index: dict) -> tuple[dict | None, float, str, str]:
    """
    Find best Yahoo category match for a Rakuten genre.
    Returns (yahoo_row, score, confidence, reasoning).
    """
    parts = parse_rakuten_genre(rakuten_genre_name)
    leaf = normalize(parts[-1])
    rakuten_top = parts[0]
    domain_keywords = get_yahoo_domain_keywords(rakuten_top)

    candidates = []

    # Strategy 1: Exact name match
    if leaf in yahoo_name_index:
        for yc in yahoo_name_index[leaf]:
            s, r = score_match(parts, yc, domain_keywords)
            candidates.append((s, yc, r))

    # Strategy 2: Partial name match (leaf in yahoo name or vice versa)
    for name_key, yc_list in yahoo_name_index.items():
        if name_key == leaf:
            continue
        if len(leaf) >= 2 and leaf in name_key:
            for yc in yc_list:
                s, r = score_match(parts, yc, domain_keywords)
                candidates.append((s, yc, r))
        elif len(name_key) >= 2 and name_key in leaf:
            for yc in yc_list:
                s, r = score_match(parts, yc, domain_keywords)
                candidates.append((s, yc, r))

    # Strategy 3: Search by path content if few good candidates
    good_candidates = [c for c in candidates if c[0] >= 50]
    if len(good_candidates) < 2 and len(leaf) >= 2:
        for yc in yahoo_cats:
            if leaf in yc["_path_norm"]:
                s, r = score_match(parts, yc, domain_keywords)
                candidates.append((s, yc, r))

    # Deduplicate by yahoo id, keeping highest score per id
    best_by_id: dict[str, tuple[float, dict, str]] = {}
    for s, yc, r in candidates:
        yid = yc["id"]
        if yid not in best_by_id or s > best_by_id[yid][0]:
            best_by_id[yid] = (s, yc, r)

    if not best_by_id:
        return None, 0, "要手動確認", "候補なし"

    # Sort by score descending
    sorted_candidates = sorted(best_by_id.values(), key=lambda x: -x[0])
    best_score, best_yc, best_reason = sorted_candidates[0]

    # Confidence
    if best_score >= 100:
        confidence = "高"
    elif best_score >= 50:
        confidence = "中"
    elif best_score >= 30:
        confidence = "低"
    else:
        return None, best_score, "要手動確認", f"スコア低({best_score:.0f}): {best_reason}"

    return best_yc, best_score, confidence, best_reason


def build_indexes(yahoo_cats: list[dict]):
    """Build name-based indexes for fast lookup."""
    name_index: dict[str, list[dict]] = defaultdict(list)
    seen = set()
    for yc in yahoo_cats:
        n = yc["_name_norm"]
        key = (n, yc["id"])
        if key not in seen:
            seen.add(key)
            name_index[n].append(yc)
    return name_index


def write_excel(rows: list[dict], path: Path):
    """Write formatted Excel file with color coding."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxlが見つかりません。Excel出力をスキップします。")
        print("  pip install openpyxl でインストールしてください。")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "カテゴリマッピング"

    # Headers
    headers = [
        "楽天ジャンルID", "楽天ジャンル名", "YahooカテゴリID", "Yahooカテゴリ名",
        "Yahooパス例", "商品数", "ステータス", "信頼度"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    orange_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for i, row in enumerate(rows, 2):
        values = [
            row.get("rakuten_genre_id", ""),
            row.get("rakuten_genre_name", ""),
            row.get("yahoo_category_id", ""),
            row.get("yahoo_category_name", ""),
            row.get("yahoo_path_example", ""),
            row.get("product_count", ""),
            row.get("status", ""),
            row.get("confidence", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [2, 5]))

            status = row.get("status", "")
            if status == "推定":
                cell.fill = orange_fill
            elif status == "要手動確認":
                cell.fill = yellow_fill

    # Column widths
    col_widths = [15, 50, 15, 25, 45, 10, 12, 8]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze panes and auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows) + 1}"

    wb.save(path)
    print(f"Excel出力: {path}")


def main():
    print("=" * 70)
    print("楽天→Yahoo カテゴリ自動推定マッピング")
    print("=" * 70)

    # Load data
    print(f"\nYahooカテゴリマスタ読込中: {YAHOO_CSV}")
    yahoo_cats = load_yahoo_categories(YAHOO_CSV)
    print(f"  → {len(yahoo_cats)} 行")

    print(f"マッピングCSV読込中: {MAPPING_CSV}")
    mapping = load_mapping(MAPPING_CSV)
    print(f"  → {len(mapping)} 行")

    matched = [r for r in mapping if r.get("status") == "マッチ済み"]
    unmatched = [r for r in mapping if r.get("status") != "マッチ済み"]
    print(f"  マッチ済み: {len(matched)} 行")
    print(f"  未マッチ:   {len(unmatched)} 行")

    # Build indexes
    name_index = build_indexes(yahoo_cats)

    # Process unmatched rows
    print(f"\n{'─' * 70}")
    print("マッチング開始")
    print(f"{'─' * 70}\n")

    stats = {"推定": 0, "要手動確認": 0, "高": 0, "中": 0, "低": 0}

    for row in mapping:
        if row.get("status") == "マッチ済み":
            row["confidence"] = ""
            continue

        genre = row["rakuten_genre_name"]

        # Check manual overrides first
        if genre in MANUAL_OVERRIDES:
            override = MANUAL_OVERRIDES[genre]
            if override is not None:
                yid, yname, ypath = override
                row["yahoo_category_id"] = yid
                row["yahoo_category_name"] = yname
                row["yahoo_path_example"] = ypath
                row["status"] = "推定"
                row["confidence"] = "高"
                stats["推定"] += 1
                stats["高"] += 1
                print(f"[高/手動] {genre}")
                print(f"  → {yname} (id={yid})")
                print(f"    パス: {ypath}")
                print(f"    理由: 手動オーバーライド")
                print()
                continue

        # Auto-match
        best, score, confidence, reasoning = find_best_match(
            genre, yahoo_cats, name_index
        )

        if best and confidence != "要手動確認":
            row["yahoo_category_id"] = best["id"]
            row["yahoo_category_name"] = best["name"]
            row["yahoo_path_example"] = best["path_name"]
            row["status"] = "推定"
            row["confidence"] = confidence
            stats["推定"] += 1
            stats[confidence] += 1
            print(f"[{confidence}] {genre}")
            print(f"  → {best['name']} (id={best['id']}, score={score:.0f})")
            print(f"    パス: {best['path_name']}")
            print(f"    理由: {reasoning}")
            print()
        else:
            row["status"] = "要手動確認"
            row["confidence"] = ""
            stats["要手動確認"] += 1
            print(f"[要手動確認] {genre}")
            print(f"    理由: {reasoning}")
            print()

    # Summary
    print(f"\n{'=' * 70}")
    print("結果サマリ")
    print(f"{'=' * 70}")
    print(f"  マッチ済み (既存):  {len(matched)} 行")
    print(f"  推定 (今回):        {stats['推定']} 行")
    print(f"    信頼度 高:        {stats['高']} 行")
    print(f"    信頼度 中:        {stats['中']} 行")
    print(f"    信頼度 低:        {stats['低']} 行")
    print(f"  要手動確認:         {stats['要手動確認']} 行")

    # Write CSV
    fieldnames = [
        "rakuten_genre_id", "rakuten_genre_name", "yahoo_category_id",
        "yahoo_category_name", "yahoo_path_example", "product_count",
        "status", "confidence"
    ]
    with open(OUTPUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(mapping)
    print(f"\nCSV出力: {OUTPUT_CSV}")

    # Write Excel
    write_excel(mapping, OUTPUT_XLSX)

    print("\n完了!")


if __name__ == "__main__":
    main()
