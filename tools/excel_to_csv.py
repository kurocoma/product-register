#!/usr/bin/env python
"""Excel 商品管理シート.xlsm の対象シートを UTF-8 CSV に変換する（統合マスタ取込用）。

出力先: docs/ネクストエンジン/from-excel/excel_*.csv
将来の「NE/Excel 自動DLツール」はこの出力を置換するだけで統合マスタ取込APIに連携できる。

使い方:
  python tools/excel_to_csv.py [xlsm_path]
  （省略時は既定の 商品管理シート.xlsm パスを使う）
"""
import csv
import sys
from pathlib import Path

import openpyxl

DEFAULT_XLSM = r"C:\Users\hppym\株式会社しまのや\くりまポータル - ドキュメント\商品管理シート.xlsm"

# シート名 → 出力ファイル名
SHEETS = {
    "商品マスタ": "excel_syohin_master.csv",
    "終売品マスタ": "excel_discon.csv",
    "商品コード一覧楽天": "excel_mall_rakuten.csv",
    "商品コード一覧Yahoo": "excel_mall_yahoo.csv",
    "商品コード一覧amazon": "excel_mall_amazon.csv",
    "しまのや商品コード一覧": "excel_mall_shimanoya.csv",
}


def main() -> int:
    xlsm = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(DEFAULT_XLSM)
    if not xlsm.exists():
        print(f"xlsm が見つかりません: {xlsm}")
        return 1

    out_dir = Path(__file__).resolve().parent.parent / "docs" / "ネクストエンジン" / "from-excel"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    for sheet_name, out_name in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  (skip) シート無し: {sheet_name}")
            continue
        ws = wb[sheet_name]
        out_path = out_dir / out_name
        rows = 0
        with open(out_path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            for r in ws.iter_rows(values_only=True):
                # 末尾の None セルを除去しつつ、各セルを文字列化（None→空）
                cells = ["" if c is None else str(c) for c in r]
                while cells and cells[-1] == "":
                    cells.pop()
                if not cells:
                    continue
                w.writerow(cells)
                rows += 1
        print(f"  {sheet_name} -> {out_path.name}  ({rows} 行)")
    wb.close()
    print(f"完了: {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
